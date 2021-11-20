import os, win32com.client
import pandas as pd
from datetime import datetime
import time
from tqdm import trange
import logging
import win32com.client as win32

main_path = r"C:\Users\User01\Desktop\EIS\Cal Poly\EIS Team - MSMT_data"



summary_folder = os.path.join(r"C:\Users\User01\Desktop\EIS\Cal Poly\EIS Team - MSMT_data", '75kOhm')
sub_folder = os.path.join(r"C:\Users\User01\Desktop\EIS\Cal Poly\EIS Team - MSMT_data", '75kOhm', 'Inidividual_test_result')
# writer = pd.ExcelWriter(os.path.join(summary_folder, 'Summary.xlsx'))

from openpyxl import load_workbook

# with pd.ExcelWriter(os.path.join(summary_folder, 'compiled.xlsx'), engine='openpyxl') as writer:
#     pd.DataFrame({}).to_excel(writer, sheet_name='Sheet1', index=False)
#     writer.save()

def concat_ws(parent, folder, bool):
    """concatenate all workbooks under folder into single workbook
    parent: parent file path
    folder: folder path where the individual worksheets are
    """
    compiled_wb = load_workbook(parent)
    for root, dirs, files in os.walk(folder):
        for file in files:
            if ('Summary' in file) == bool:
                wb = load_workbook(filename = os.path.join(folder, file)).active
                wb._parent = compiled_wb
                compiled_wb._add_sheet(wb)
    compiled_wb.save(parent)


def summary(folder_path):
    import warnings
    warnings.filterwarnings('ignore', category=UserWarning, module='xlsxwriter')
    """generate summary file, overlay all the results in one plot
    output_dir: where the summary file is saved
    input_dir: where the data files are
    """
    folder_name = os.path.basename(folder_path)
    output_path = os.path.join(folder_path, f'{folder_name}Summary.xlsx')
    Summary = pd.ExcelWriter(output_path)
    summary_table = []
    sheetnames = []
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if 'Summary' not in file:
                summary_table.append(pd.read_excel(os.path.join(folder_path, file), nrows=1, usecols=list(range(9,21))))
                sheetnames.append(file.split('.xlsx')[0]) 
        df = pd.concat(summary_table)
        time_elapsed = df['Time Elapsed [s]'].to_list()
        df['date'] = pd.to_datetime(df['date'])  # convert date column in datetime format
        end_row = df['Samples'].values[0]
        df.sort_values(by=['date'], ascending=False, inplace=True)
        df.to_excel(Summary, f'{folder_name}Summary', index=False,)
        worksheet = Summary.sheets[f'{folder_name}Summary']
        # Link to another Excel workbook.
        row = 0
        for file in files:
            if 'Summary' not in file:
                worksheet.write_url(row+1, df.shape[1]-1, f"internal:'{file.split('.xlsx')[0]}'!A1", string=str(time_elapsed[row]))
                row += 1
        worksheet.set_column(0, df.shape[1]-1, 17.5)
        workbook = Summary.book
        ## overlay plots
        
        for col, label in enumerate(['Impedance [kOhm]', 'Resistance [kOhm]', 'Reactance [kOhm]', 'Phase [degree]', 'Series Capaitance [uF]', 'Parallel Capaitance [nF]', 'Series Inductance [H]', 'Parallel Inductance [H]']):
            chart = workbook.add_chart({'type': 'scatter'})
            for sheet_num, sheetname in enumerate(sheetnames):    
                chart.add_series({
                    'name': '{:.3f}'.format(time_elapsed[sheet_num]),
                    'categories': [sheetname, 1, 0, end_row, 0],  # x-axis
                    'values': [sheetname, 1, col+1, end_row, col+1],    # y-axis
                    'line': {'none': True},
                    'marker': {'type': 'circle', 'size': 3}
                })
                chart.set_x_axis({
                    'name': 'Frequency [kHz]',
                    'major_gridline': {'visible': True},
                    'log_base': 10,
                    'min': 1e3, 'max': 10e6,
                    'display_units': 'thousands', 
                    'display_units_visible': False,
                    'minor_tick_mark': 'cross'})
                if label == 'Phase [degree]':
                    chart.set_y_axis({'name': label, 'major_gridline': {'visible': True}, 'major_unit': 30})
                elif 'kOhm' in label:
                    chart.set_y_axis({
                        'name': label,
                        'major_gridline': {'visible': True},
                        'display_units': 'thousands', 
                        'display_units_visible': False,
                        })
                else:
                    chart.set_y_axis({'name': label, 'major_gridline': {'visible': True}})
                chart.set_title({'name': label})
                chart.set_legend({'position': 'bottom'})
                chart.set_size({'x_scale': 1.2, 'y_scale': 1.5})
            worksheet.insert_chart(3+col*22, df.shape[1], chart)        # df.shape[0]: len(df.index), df.shape[1]: len(df.columns)

        Summary.save()
        concat_ws(output_path, folder_path, False) # concatenate all workbooks under folder into single workbook

print("Summary" in '10_24_15_07_2021Summary.xlsx')
summary(r'C:\Users\User01\Desktop\EIS\Cal Poly\EIS Team - MSMT_data\75kOhm\Inidividual_test_result\10_24_15_07_2021')

# for folder in os.scandir(sub_folder):
#     summary(folder)

# for root, dirs, files in os.walk(sub_folder):
#         for file in files:
#             print(file)

# concat_ws(r'C:\Users\User01\Desktop\EIS\Cal Poly\EIS Team - MSMT_data\75kOhm\Summary.xlsx', r'C:\Users\User01\Desktop\EIS\Cal Poly\EIS Team - MSMT_data\75kOhm\Inidividual_test_result\10_24_15_07_2021', True)
# import openpyxl
# import contextlib

# with contextlib.suppress(IOError), open('file') as f:
#     pass
# wb = load_workbook(r'C:\Users\User01\Desktop\EIS\Cal Poly\EIS Team - MSMT_data\75kOhm\Inidividual_test_result\12_42_2021_10_24\Summary_12_42_2021_10_24.xlsx')

# for sheet in wb.worksheets[1:]:
#     wb.remove(sheet)
# wb2 = load_workbook(r'C:\Users\User01\Desktop\EIS\Cal Poly\EIS Team - MSMT_data\75kOhm\Inidividual_test_result\10_24_15_07_2021\Summary_10_24_15_07_2021.xlsx')
# for ws in wb:
#     ws._parent = wb
#     wb._add_sheet(ws)
# wb.save(r'C:\Users\User01\Desktop\EIS\Cal Poly\EIS Team - MSMT_data\75kOhm\compiled.xlsx')