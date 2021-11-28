import warnings
from openpyxl import load_workbook
import os
import pandas as pd

def concat_ws(parent, folder, bool):
    """concatenate all workbooks under folder into single workbook
    parent: parent file path
    folder: folder path where the individual worksheets are
    """
    compiled_wb = load_workbook(parent)
    for root, dirs, files in os.walk(folder):
        for file in files:
            if ('Summary' in file) == bool:
                wb = load_workbook(filename = os.path.join(folder, file)).active
                wb._parent = compiled_wb
                compiled_wb._add_sheet(wb)
    compiled_wb.save(parent)


def summary(folder_path):
    """generate summary file, overlay all the results in one plot
    output_dir: where the summary file is saved
    input_dir: where the data files are
    """
    warnings.filterwarnings('ignore', category=UserWarning, module='xlsxwriter')
    folder_name = os.path.basename(folder_path)
    output_path = os.path.join(folder_path, f'{folder_name}Summary.xlsx')
    Summary = pd.ExcelWriter(output_path)
    summary_table = []
    sheetnames = []
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if 'Summary' not in file:
                summary_table.append(pd.read_excel(os.path.join(folder_path, file), nrows=1, usecols=list(range(9,21))))
                sheetnames.append(file.split('.xlsx')[0]) 
        df = pd.concat(summary_table)
        time_elapsed = df['Time Elapsed [s]'].to_list()
        df['date'] = pd.to_datetime(df['date'])  # convert date column in datetime format
        end_row = df['Samples'].values[0]
        df.sort_values(by=['date'], ascending=False, inplace=True)
        df.to_excel(Summary, f'{folder_name}Summary', index=False,)
        worksheet = Summary.sheets[f'{folder_name}Summary']
        # Link to another Excel workbook.
        row = 0
        for file in files:
            if 'Summary' not in file:
                worksheet.write_url(row+1, df.shape[1]-1, f"internal:'{file.split('.xlsx')[0]}'!A1", string=str(time_elapsed[row]))
                row += 1
        worksheet.set_column(0, df.shape[1]-1, 17.5)
        workbook = Summary.book
        ## overlay plots
        
        for col, label in enumerate(['Impedance [kOhm]', 'Resistance [kOhm]', 'Reactance [kOhm]', 'Phase [degree]', 'Series Capaitance [uF]', 'Parallel Capaitance [nF]', 'Series Inductance [H]', 'Parallel Inductance [H]']):
            chart = workbook.add_chart({'type': 'scatter'})
            for sheet_num, sheetname in enumerate(sheetnames):    
                plot_chart(chart, '{:.3f}'.format(time_elapsed[sheet_num]), sheetname, end_row, col, label, 1e3, 10e6)
            worksheet.insert_chart(3+col*22, df.shape[1], chart)        # df.shape[0]: len(df.index), df.shape[1]: len(df.columns)

        Summary.save()
        concat_ws(output_path, folder_path, False) # concatenate all workbooks under folder into single workbook

def plot_chart(chart, name, sheetname, endrow, col, label, min_, max_):
    chart.add_series({
        'name': name,
        'categories': [sheetname, 1, 0, endrow, 0],  # x-axis
        'values': [sheetname, 1, col+1, endrow, col+1],    # y-axis
        'line': {'none': True},
        'marker': {'type': 'circle', 'size': 3}
    })
    chart.set_x_axis({
        'name': 'Frequency [kHz]',
        'major_gridline': {'visible': True},
        'log_base': 10,
        'min': min_, 'max': max_,
        'display_units': 'thousands', 
        'display_units_visible': False,
        'minor_tick_mark': 'cross'})
    if label == 'Phase [degree]':
        chart.set_y_axis({'name': label, 'major_gridline': {'visible': True}, 'major_unit': 30})
    elif 'kOhm' in label:
        chart.set_y_axis({
            'name': label,
            'major_gridline': {'visible': True},
            'display_units': 'thousands', 
            'display_units_visible': False,
            })
    else:
        chart.set_y_axis({'name': label, 'major_gridline': {'visible': True}})
    chart.set_title({'name': label})
    chart.set_legend({'position': 'bottom'})
    chart.set_size({'x_scale': 1.2, 'y_scale': 1.5})
